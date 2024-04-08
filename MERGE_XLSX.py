
import os
import shutil
import datetime
import hashlib
import pandas as pd
import openpyxl as pyxl
from openpyxl.styles import (Alignment, Font)
from string import (ascii_uppercase as uppers,
                    punctuation as punc,
                    digits as digs)


# Подсказки
tips = {
    1: '''
Cкопируйте папки с файлами, которые вы хотите объединить
в каталог: 'DROP_FOLDER'
и нажмите ENTER\n
''',
    2: '''
Введите номер нужного каталога (только цифра)
и нажмите ENTER\n
''',
    4: '\nПопробуем снова\n',
    5: '''
Введите номер листа для сшивания файлов (только цифра)
Если нужного листа нет в списке - введите R
и нажмите ENTER\n
''',
    6: '''
Попробуем найти заголовок вручную?
введите номер строки с заголовком файлов
и нажмите ENTER\n
''',
    7: '''
Если заголовок некорректный, введите R
и нажмите ENTER
Если результат устраивает,
нажмите ENTER без ввода символов\n
''',
    8: '''
Введите номер столбца с информацией
о порядковом номере (№)
нажмите ENTER \n
''',
    10: '''
Если вы уверены, что найденные ошибки - это опечатки
и стрктуры таблиц сохранены:
нажмите ENTER без ввода символов\n
В ином случае, рекомендуется отредактировать указанные файлы
и запустить повторную проверку вводом комманды R
программу при этом можно не закрывать\n
НАСТОЯТЕЛЬНО НЕ РЕКОМЕНДУЕТСЯ ИГНОРИРОВАТЬ ПРЕДУПРЕЖДЕНИЯ О:
    --- Отсутствии целевого листа;
    --- Разной длинне заголовков;
    --- Совпадении заголовка менее 90%;
    --- Несовпадающем порядке столбцов (если не обнаружено отличающихся ячеек)
'''}


# Переменные
droper_folder = 'DROP_FOLDER'
result_folder = 'RESULT'
temp_folder = 'temp'
let_set = dict(enumerate(list(uppers) + ['A' + let for let in uppers], 1))
p_set = punc + ''
pn_set = punc + digs
null_values = ['0', 'none', 'None', 'null', ' ', 'nan', '']
out_sheet_title = 'Output'
r_limiter = (0, 0)
actions = ['r', 'e', '']
header_auto_set = True
numerator_in = True
numerator_auto_set = True
numerator_col = 0

# Извлечение названий доступных файлов из целевого каталога:


def get_filelist(d_dir, t_dir):
    file_list = list(os.walk(f'{d_dir}/{t_dir}'))[0][2]
    return file_list

# Извлечение хеш суммы файла:


def get_f_hash(d_dir, t_dir, f_name):
    with open(f'{d_dir}/{t_dir}/{f_name}', 'rb') as f:
        f_hash = hashlib.md5()
        while chunk := f.read(8192):
            f_hash.update(chunk)
    return f_hash.hexdigest()


# Упаковщик в селекторы


def pack_my_list(lst):
    dic = {key: val for key, val in enumerate(lst, 1)}
    [print(f'№{key}  ---  {val}') for key, val in dic.items()]
    return dic

# Проверка ввода, критерий - список числовых значений, tip - подсказка


def cheсk_input(tip, criteria=[]):
    repeat_check = True
    while repeat_check:
        feedback = input(tip).lower().strip(p_set)
        if feedback in actions:
            return feedback
        elif feedback.isdigit():
            if int(feedback) in criteria:
                return int(feedback)
            else:
                print('\nТакого поля нет в диапазоне\n')
        else:
            print('''
Недопустимое значение :(
Пожалуйста, повторите ввод
Корректные значения:
    --- Пустая строка (если указано)
    --- Целое число
    --- Значение из предложенного диапазона
    --- Комманда

Список комманд:
    --- Возврат к предыдущему шагу - R
    --- Завершение процесса - E \n
        ''')


# Забрать датасет из файла
def get_df(d_dir, t_dir, f_name, t_sheet):
    wb = pyxl.load_workbook(f'{d_dir}/{t_dir}/{f_name}',
                            read_only=True,
                            data_only=True)
    if t_sheet in wb.sheetnames:
        ws = wb[t_sheet]
        df = pd.DataFrame(ws.values)
    else:
        df = pd.DataFrame()
    wb.close()
    return df


# Поиск заголовка на листе - На выход позиция строки в датасете:
def find_my_header(df):
    # Срез 20 первых строк, если размер датафрейма больше 20
    if df.shape[0] > 20:
        df_slice = df.iloc[:20].copy()
    else:
        df_slice = df.copy()
    # Удаление полностью пустых строк
    df_slice.dropna(axis=0,
                    how='all',
                    inplace=True)
    all_rows = {}
    for row in df_slice.itertuples():
        # Исключение пустых ячеек из строки
        iter_row = tuple(filter(lambda val: str(val).lower() not in null_values, row[1:]))
        # Количество ячеек содержащих только буквы
        all_rows[row[0]] = sum(map(lambda val:
                                   len(str(sorted(str(val))).strip(pn_set)) != 0,
                                   iter_row))
    return max(all_rows.items(), key=lambda val: val[1])[0]


# Преобразование заголовка - на вход iloc:
def list_header(row):
    new_row = list(row)
    empty_test = len(
        list(filter(lambda val: str(val).lower() not in null_values, new_row)))
    if empty_test > 3:
        while str(new_row[-1]).lower() in null_values and len(new_row) > 0:
            new_row.pop()
        new_row = list(map(lambda val: ' '.join(str(val).lower().split()), new_row))
        new_row = list(map(lambda val: '' if val in null_values else val, new_row))
        return new_row
    else:
        return ['Выбранная строка пуста или содержит менее 3 ячеек с данными']


# Показать прогресс проверок
def print_progress(i, total):
    print(f'\n{round(i / total * 100,1)}%')


# Сравнение эталона заголовка и заголовка следующего файла:
def headers_seq(ref_h, next_h):
    # Наличие листа (по умолчанию - 1)
    # Список отличий
    h_diff = set(ref_h) ^ set(next_h)
    # Кол-во отличий
    h_len_dif = abs(len(next_h) - len(ref_h))
    # % совпадения
    match = round((len(ref_h) - len(h_diff))/len(ref_h) * 100, 1)
    # Совпадение порядка ячеек
    col_seq = ref_h == next_h
    return (1, h_diff, h_len_dif, match, col_seq)


# Показать лог найденных ошибок
def print_log(l_dict, act_files):
    if len(l_dict) == 0:
        print('\nВсё в порядке - ошибок не обнаружено :)\n')
        return ('\nНажмите ENTER без ввода символов, чтобы начать запись в файл\n')
    else:
        print('\nЯ нашел парочку ошибок:\n')
        for key, val in l_dict.items():
            if key in act_files:
                print(key)
                if val[0] == 0:
                    print('--- Целевой лист не найден')
                else:
                    seq_stat = ['отличается', 'совпадает']
                    print(f'--- Различия заголовка: - {val[1]}',
                          f'--- Различия в длинне заголовка: - {val[2]} столбцов',
                          f'--- Заголовки совпадают на: - {val[3]}%',
                          f'--- Порядок ячеек: - {seq_stat[val[4]]}',
                          sep='\n')
                print('\n')
        return tips[10]


# Проверка доступности файлов

def check_filelist(d_dir, t_dir, files_dict):
    act_f_lst = set(get_filelist(d_dir, t_dir))
    cur_f_lst = set(files_dict.keys())
    del_files = cur_f_lst - act_f_lst
    new_files = act_f_lst - cur_f_lst
    if len(new_files) == len(del_files) == 0:
        return act_f_lst
    else:
        if len(new_files) > 0:
            for file in new_files:
                files_dict[file] = [0, 0]
        if len(del_files) > 0:
            [files_dict.pop(file) for file in del_files]
        return act_f_lst

# проверка на добавление с список исключений


def ban_hammer(errors_dict, f_name):
    if f_name in errors_dict.keys():
        log = errors_dict[f_name]
        return not (log[2] == 0 and log[3] > 70)
    else:
        return False

# ПРОВЕРИТЬ!
# Проверка файлов и запись в temp


def check_files(d_dir,
                t_dir,
                temp_dir,
                t_sheet,
                ref_head,
                files_dict,
                errors_dict,
                auto_find_h,
                h_pos,
                num_in,
                num_auto,
                num_pos):
    banned_files = []
    act_f_list = check_filelist(d_dir, t_dir, files_dict)
    print(f'\nКаталог:  --- {t_dir}\nПроверка: --- {len(act_f_list)} файлов\n')
    for num, file in enumerate(list(act_f_list), 1):
        cur_hash = get_f_hash(d_dir, t_dir, file)
        # Если хеш файла не совпадает с записанным
        #print(cur_hash, files_dict[file])
        if cur_hash not in files_dict[file]:
            df = get_df(d_dir, t_dir, file, t_sheet)
            files_dict[file][0] = cur_hash
            # Логировать файл не содержащий целевой лист
            if df.empty:
                errors_dict[file] = 0, 0, 0, 0, 0
                # Если ошибок в листах нет - инициировать проверку заголовков
            else:
                if auto_find_h:
                    targ_row = find_my_header(df)
                    next_header = list_header(df.iloc[targ_row])
                    files_dict[file][1] = targ_row
                else:
                    next_header = list_header(df.iloc[h_pos])
                    files_dict[file][1] = h_pos
                check_header = headers_seq(ref_head, next_header)
                if check_header[4] == 0 or len(check_header[1]) > 0:
                    errors_dict[file] = check_header
                else:
                    errors_dict.setdefault(file)
                    errors_dict.pop(file)
            if ban_hammer(errors_dict, file):
                banned_files.append(file)
            else:
                df = extract_clear_data(df,
                                        file,
                                        files_dict[file][1],
                                        len(ref_head),
                                        num_in,
                                        num_auto,
                                        num_pos)
                df.to_pickle(f'{temp_dir}/{file}.pkl')
        print_progress(num, len(act_f_list))
    return act_f_list - set(banned_files)

# Создание файла для записи


def create_res_file(path, file_name, sheet_name, header):
    wb = pyxl.Workbook()
    ws = wb.active
    ws.append(['исходный файл'] + header)
    ws.title = sheet_name  # константа
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 60
    for col, n in zip(let_set.values(), range(len(header)+1)):
        ws.column_dimensions[col].width = 25
        ws[col+'1'].font = Font(bold=True, name='Times New Roman', size=12)
    for cell in ws[1]:
        if cell.value:
            cell.alignment = Alignment(
                horizontal='left', vertical='center', wrap_text=True)
    wb.save(f'{path}/{file_name}.xlsx')
    wb.close()
    return f'{path}/{file_name}.xlsx'


# Поиск нумератора - возвращает кортеж (нумератор найден, позиция столбца)


def find_numerator(df):
    if df.shape[0] > 11:
        df_slice = df.iloc[:11].transpose().copy()
    else:
        df_slice = df.copy()
    cols = {}
    # Критерий - числовое значение, каждое последующее значение больше пр. на 1, 2 или 3
    for row in df_slice.itertuples():
        try:
            cols[row[0]] = sum(map(lambda n1, n2: (str(n1) + str(n2)).isdigit() and n2
                                   in [int(n1)+1, int(n1)+2, int(n1)+3],
                                   row[1:], row[2:]))
        except:
            pass
    # Больше половины значений соответствует критерию
    cols = dict(filter(lambda item: item[1] > 5, cols.items()))
    if len(cols) == 0:
        return (False, 0)
    else:
        targ_col = max(cols.items(), key=lambda val: val[1])[0]
        return (True, targ_col)

# Удаление из хвоста висячих строк


def clear_tale(df_tail):
    drop_list = []
    for row in df_tail.itertuples():
        # Количество ненулевых ячеек
        res = sum(map(lambda val: str(val) not in null_values, row[1:]))
        if res < 3:
            drop_list.append(row[0])
    return drop_list

# Преобразование и очистка датасета


def extract_clear_data(dirty_df,
                       file_ident,
                       h_pos,
                       cols_to_df,
                       num_in=True,
                       num_auto=True,
                       num_pos=0):
    df = dirty_df.iloc[h_pos+1:, :cols_to_df]
    df.reset_index(drop=True, inplace=True)
    if num_in:
        if num_auto:
            numerator = find_numerator(df)
            if numerator[0]:
                num_pos = numerator[1]
                df.dropna(subset=df.columns.difference([num_pos]),
                          how='all',
                          inplace=True)
                df.drop_duplicates(subset=df.columns.difference([num_pos]),
                          keep='last',
                          inplace=True,
                          ignore_index=True)
            else:
                df.dropna(how='all',
                          inplace=True)
                df.drop_duplicates(keep='last',
                                   inplace=True)
        else:
                df.dropna(subset=df.columns.difference([num_pos]),
                          how='all',
                          inplace=True)
                df.drop_duplicates(subset=df.columns.difference([num_pos]),
                          keep='last',
                          inplace=True,
                          ignore_index=True)
    else:
        df.dropna(axis=0,
                  how='all',
                  inplace=True)
        df.drop_duplicates(keep='last', inplace=True)
    print(numerator)
    # Удаление пустого хвоста
    if df.shape[0] > 3:
        drop_list = clear_tale(df.tail(3))
        if len(drop_list) != 0:
            df.drop(labels=drop_list, axis=0, inplace=True)
    # Добавление названия источника датасета
    df.insert(0, 'file_ident', file_ident.rstrip('.xlsx'))
    return df


# Запись в файл построчно из листа с df.pickle

def write_to_res(path, write_list, target_sheet):
    wb = pyxl.load_workbook(path, data_only=True)
    ws = wb[target_sheet]
    for file in write_list:
        df = pd.read_pickle(f'temp/{file}.pkl')
        for row in df.itertuples():
            ws.append(row[1:])
        wb.save(path)
    wb.close


# START
input(tips[1])


repeat = True
while repeat:
    try:
        os.makedirs(temp_folder, exist_ok=True)
        os.makedirs(result_folder, exist_ok=True)
    except:
        pass
    
    time1 = datetime.datetime.now()
    
    # Поиск доступных для обработки каталогов с файлами
    folder_list = list(os.walk(droper_folder))[0][1]
    print("\nДоступные каталоги:\n")
    folder_list = pack_my_list(folder_list)

    # Выбор каталога для обработки
    feedback = cheсk_input(tips[2], folder_list.keys())
    if feedback in ['', 'r']:
        print(tips[4])
        continue
    elif feedback in ['e']:
        exit()

    # Присвоение выбранного значения переменной - целевая папка
    targ_dir = folder_list[feedback]

    # Извлечение списка файлов из каталога
    start_files_list = get_filelist(droper_folder, targ_dir)

    # Листы доступные в первом файле выбранного каталога
    wb = pyxl.load_workbook(
        f'{droper_folder}/{targ_dir}/{start_files_list[0]}', read_only=True).sheetnames
    print("Эти листы я нашел в первом файле:\n")
    sheet_list = pack_my_list(wb)

    # Выбор целевого листа
    feedback = cheсk_input(tips[5], sheet_list.keys())
    if feedback in ['', 'r']:
        print(tips[4])
        continue
    elif feedback in ['e']:
        exit()

    # Присвоение выбранного значения переменной - целевой лист
    targ_sheet = sheet_list[feedback]

    # Достаем датасет из целевого листа первого файла
    df = get_df(droper_folder, targ_dir, start_files_list[0], targ_sheet)

    # Инициируем поиск и выбор заголовка
    repeat_header_set = True
    while repeat_header_set:

        # Выбор позиции заголовка автоматически
        if header_auto_set:
            targ_row = find_my_header(df)

        # Ручной выбор позиции заголовка
        else:
            feedback = cheсk_input(tips[6], range(1, 40))
            if feedback in ['', 'r']:
                print(tips[4])
                continue
            elif feedback in ['e']:
                exit()
            targ_row = feedback - 1

        # Вывод заголовка
        ref_header = list_header(df.iloc[targ_row])
        header_dict = pack_my_list(ref_header)

        # Подтверждение заголовка
        feedback = cheсk_input(tips[7], header_dict.keys())
        if feedback in ['r']:
            print(tips[4])
            header_auto_set = False
            continue
        elif feedback in ['e']:
            exit()
        else:
            repeat_header_set = False

    # Выбор позиции столбца-нумератора
    if numerator_in and not numerator_auto_set:
        feedback = cheсk_input(tips[8], range(1, len(ref_header)+1))
        if feedback in ['', 'r']:
            print(tips[4])
            continue
        elif feedback in ['e']:
            exit()
        numerator_col = feedback - 1

    # Инициализация проверки файлов на корректность
    # Создание словаря свойств файла - имя: хеш, позиция заголовка

    repeat_filecheck = True
    files_list = {file: [0, 0] for file in start_files_list}
    scan_error = {}
    while repeat_filecheck:
        lst_for_write = check_files(droper_folder,
                                    targ_dir,
                                    temp_folder,
                                    targ_sheet,
                                    ref_header,
                                    files_list,
                                    scan_error,
                                    header_auto_set,
                                    targ_row,
                                    numerator_in,
                                    numerator_auto_set,
                                    numerator_col)
        # Вывести список ошибок, если есть
        log_res = print_log(scan_error, lst_for_write)

    # Запрос на перепроверку файлов
        feedback = cheсk_input(log_res)
        if feedback in ['r']:
            print(tips[4])
            continue
        elif feedback in ['e']:
            exit()
        else:
            repeat_filecheck = False

    # Создание результирующей таблицы и запись переменной с путем файла
    last_res_path = create_res_file(result_folder, targ_dir, out_sheet_title, ref_header)

    # Запись в файл
    print(f'\nКаталог:  --- {targ_dir}\nЗапись: --- {len(lst_for_write)} файлов\n')

    write_to_res(last_res_path,
                 lst_for_write,
                 out_sheet_title)
    time2 = datetime.datetime.now()
    total_time = time2 - time1
    shutil.rmtree(temp_folder, ignore_errors=True)
    print(total_time)
